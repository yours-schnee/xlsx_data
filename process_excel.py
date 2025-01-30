import pandas as pd
import openpyxl
from openpyxl import Workbook
from pathlib import Path

def clean_sheet_name(name):
    invalid_chars = [':', '/', '\\', '?', '*', '[', ']']
    clean_name = str(name)
    for char in invalid_chars:
        clean_name = clean_name.replace(char, '_')
    return clean_name[:31]

def process_directory_excel(directory_path):
    excel_files = sorted(list(Path(directory_path).glob("*.xlsx")))
    if not excel_files:
        return None
    
    print(f"\nProcessing directory: {directory_path}")
    print(f"Found {len(excel_files)} Excel files")
    data_dict = {}
    
    for excel_file in excel_files:
        print(f"Processing file: {excel_file}")
        df = pd.read_excel(excel_file)
        
        for col_name in df.columns:
            # 列の全データを取得
            column_data = df[col_name].dropna()
            if len(column_data) < 2:  # 少なくとも2行（シート名とデータ1行）必要
                continue
                
            # 1行目をシート名として使用
            sheet_name = clean_sheet_name(str(column_data.iloc[0]))
            # 2行目以降をデータとして使用
            data_values = column_data.iloc[1:].tolist()
            
            if sheet_name not in data_dict:
                data_dict[sheet_name] = []
            data_dict[sheet_name].append([excel_file.name, data_values])
            print(f"Added data for sheet '{sheet_name}' from {excel_file.name}")
    
    return data_dict

def save_workbook(data_dict, output_file):
    if not data_dict:
        return
    
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除
    
    for sheet_name, file_data_list in data_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        print(f"Creating sheet: {sheet_name}")
        
        for col_idx, (filename, values) in enumerate(file_data_list, start=1):
            # ファイル名を1行目に配置
            ws.cell(row=1, column=col_idx, value=filename)
            
            # データを2行目以降に配置
            for row_idx, value in enumerate(values, start=2):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(output_file)
    print(f"Saved to {output_file}")

def process_multiple_excel(input_dir):
    root_path = Path(input_dir)
    subdirs = [d for d in root_path.iterdir() if d.is_dir()]
    print(f"Found {len(subdirs)} subdirectories")
    
    for subdir in subdirs:
        data_dict = process_directory_excel(subdir)
        if data_dict:
            output_file = root_path / f"{subdir.name}_output.xlsx"
            save_workbook(data_dict, output_file)

if __name__ == "__main__":
    input_directory = "excel_files"
    process_multiple_excel(input_directory)
