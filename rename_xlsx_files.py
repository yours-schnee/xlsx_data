import os

def rename_xlsx_files(root_path):
    # サブフォルダを取得
    for folder_name in os.listdir(root_path):
        folder_path = os.path.join(root_path, folder_name)
        
        # フォルダであることを確認
        if os.path.isdir(folder_path):
            # フォルダ内のファイルを取得
            for file_name in os.listdir(folder_path):
                # xlsxファイルを探す
                if file_name.endswith('.xlsx'):
                    old_file_path = os.path.join(folder_path, file_name)
                    new_file_name = f"{folder_name}.xlsx"
                    new_file_path = os.path.join(folder_path, new_file_name)
                    
                    # ファイル名を変更
                    os.rename(old_file_path, new_file_path)
                    print(f"Renamed: {file_name} -> {new_file_name}")

import os
from openpyxl import load_workbook

def split_and_rename_xlsx_files(root_path):
    for folder_name in os.listdir(root_path):
        folder_path = os.path.join(root_path, folder_name)
        
        if os.path.isdir(folder_path):
            for file_name in os.listdir(folder_path):
                if file_name.endswith('.xlsx'):
                    file_path = os.path.join(folder_path, file_name)
                    
                    # Excelファイルを読み込む
                    wb = load_workbook(file_path)
                    
                    # 各シートを個別のファイルとして保存
                    for i, sheet_name in enumerate(wb.sheetnames, 1):
                        # 新しいワークブックを作成
                        new_wb = load_workbook(file_path)
                        # 保持したいシート以外を削除
                        for s in new_wb.sheetnames:
                            if s != sheet_name:
                                del new_wb[s]
                                
                        # 新しいファイル名を生成して保存
                        new_file_name = f"{folder_name}_{i}.xlsx"
                        new_file_path = os.path.join(folder_path, new_file_name)
                        new_wb.save(new_file_path)
                        print(f"Created: {new_file_name} (from sheet: {sheet_name})")

# 使用例
root_directory = "path/to/your/root/folder"
rename_xlsx_files(root_directory)


# 使用例
root_directory = "path/to/your/root/folder"
split_and_rename_xlsx_files(root_directory)



